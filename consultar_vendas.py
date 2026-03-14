import asyncio
import aiohttp
import pandas as pd
import locale
import json
import os
import time
import numpy as np
from datetime import datetime, timezone
from pathlib import Path
from shapely.geometry import Point, shape, Polygon, MultiPolygon
from scipy.spatial import KDTree
from openpyxl import load_workbook
import random

# ---------------------
# Configurações
# ---------------------

locale.setlocale(locale.LC_ALL, 'pt_BR.UTF-8')

#Faz a consulta dentro de certo dia até o momento. Se quiser pegar o histórico para trás é só trocar na url o after por before
# URL_N5 = "https://chain-history.upland.me/v2/history/get_actions?act.name=n5&after=2025-12-16T03:00:00.000Z&sort=desc&limit=1000"
# URL_N52 = "https://chain-history.upland.me/v2/history/get_actions?act.name=n52&after=2025-12-16T03:00:00.000Z&sort=desc&limit=1000"


#Faz a consulta dentro de um intervalo de dias
# URL_N5 = "https://chain-history.upland.me/v2/history/get_actions?act.name=n5&before=2025-06-18T03:00:00.000Z&after=2025-06-14T03:00:00.000Z&sort=desc&limit=1000"
# URL_N52 = "https://chain-history.upland.me/v2/history/get_actions?act.name=n52&before=2025-06-18T03:00:00.000Z&after=2025-06-14T03:00:00.000Z&sort=desc&limit=1000"



# #Faz a consulta do momento atual para trás obedecendo o limit
URL_N5 = "https://chain-history.upland.me/v2/history/get_actions?act.name=n5&sort=desc&limit=1000"
URL_N52 = "https://chain-history.upland.me/v2/history/get_actions?act.name=n52&sort=desc&limit=1000"



URL_PROPERTY = "https://api.prod.upland.me/api/properties/{}"
URL_MATCH = "https://api.upland.me/properties/match/{}"
URL_NEIGHBORHOODS = "https://api.prod.upland.me/api/neighborhood"

# Headers rotacionados para evitar bloqueio
HEADERS_LIST = [
    {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/141.0.0.0 Safari/537.36",
        "Accept": "application/json",
        "Accept-Language": "en-US,en;q=0.9"
    },
    {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:146.0) Gecko/20100101 Firefox/146.0",
        "Accept": "application/json",
        "Accept-Language": "pt-BR,pt;q=0.8,en-US;q=0.5,en;q=0.3"
    },
    {
        "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/141.0.0.0 Safari/537.36",
        "Accept": "application/json",
        "Accept-Language": "en-US,en;q=0.9"
    }
]

# Reduzir paralelismo de 8 para 3 para evitar rate limit
SEM = asyncio.Semaphore(3)
CACHE_FILE = "neighborhoods_cache.json"

# Contador de requisições
request_count = 0
last_request_time = time.time()

# ---------------------
# Util: fetch JSON com retry e rate limiting
# ---------------------
async def fetch_json(session, url, max_retries=3):
    global request_count, last_request_time
    
    async with SEM:
        # Rate limiting: máximo 10 req/segundo
        current_time = time.time()
        if current_time - last_request_time < 0.1:
            await asyncio.sleep(0.1)
        last_request_time = time.time()
        
        # Rotaciona headers
        headers = random.choice(HEADERS_LIST)
        
        for attempt in range(max_retries):
            try:
                request_count += 1
                
                async with session.get(url, headers=headers, timeout=30) as resp:
                    # Se der 429, espera mais tempo
                    if resp.status == 429:
                        wait_time = (attempt + 1) * 5  # 5s, 10s, 15s
                        print(f"⚠️ Erro 429 (Too Many Requests). Aguardando {wait_time}s...")
                        await asyncio.sleep(wait_time)
                        continue
                    
                    if resp.status != 200:
                        print(f"⚠️ Erro {resp.status} ao acessar {url}")
                        if attempt < max_retries - 1:
                            await asyncio.sleep(2 * (attempt + 1))
                            continue
                        return None
                    
                    return await resp.json()
                    
            except asyncio.TimeoutError:
                print(f"⏱️ Timeout na tentativa {attempt + 1}/{max_retries}")
                if attempt < max_retries - 1:
                    await asyncio.sleep(3 * (attempt + 1))
                    continue
                return None
            except Exception as e:
                print(f"❌ Exception ao acessar {url}: {e}")
                if attempt < max_retries - 1:
                    await asyncio.sleep(2 * (attempt + 1))
                    continue
                return None
        
        return None

# ---------------------
# Carregar / salvar cache de bairros
# ---------------------
def load_neighborhood_cache():
    if os.path.exists(CACHE_FILE):
        try:
            with open(CACHE_FILE, "r", encoding="utf-8") as f:
                cache = json.load(f)
            print("✓ Cache de bairros carregado com sucesso.")
            return cache
        except Exception as e:
            print(f"⚠️ Falha ao ler cache ({CACHE_FILE}): {e}")
    
    # Baixar agora
    try:
        import requests
        print("⏳ Baixando dados de bairros...")
        resp = requests.get(URL_NEIGHBORHOODS, headers=HEADERS_LIST[0], timeout=30)
        if resp.status_code == 200:
            data = resp.json()
            payload = {"data": data, "fetched_at": datetime.now(timezone.utc).isoformat()}
            try:
                with open(CACHE_FILE, "w", encoding="utf-8") as f:
                    json.dump(payload, f, ensure_ascii=False, indent=2)
                print(f"✓ Cache salvo em {CACHE_FILE}")
            except Exception as e:
                print(f"⚠️ Não foi possível salvar cache: {e}")
            return payload
        else:
            print(f"❌ Falha ao baixar bairros. Status: {resp.status_code}")
            return {"data": [], "fetched_at": None}
    except Exception as e:
        print(f"❌ Erro ao baixar bairros: {e}")
        return {"data": [], "fetched_at": None}

# [... resto das funções de bairro permanecem iguais ...]
def normalize_polygon_from_item(item):
    polygons = []
    boundaries = item.get("boundaries")
    if not boundaries:
        return polygons
    if isinstance(boundaries, dict):
        try:
            geom = shape(boundaries)
            if isinstance(geom, (Polygon, MultiPolygon)):
                if isinstance(geom, MultiPolygon):
                    polygons.extend(list(geom.geoms))
                else:
                    polygons.append(geom)
        except Exception:
            pass
        return polygons
    if isinstance(boundaries, str):
        try:
            parsed = json.loads(boundaries)
            geom = shape(parsed)
            if isinstance(geom, (Polygon, MultiPolygon)):
                if isinstance(geom, MultiPolygon):
                    polygons.extend(list(geom.geoms))
                else:
                    polygons.append(geom)
            return polygons
        except Exception:
            try:
                parsed = json.loads(boundaries.replace("'", '"'))
                geom = shape(parsed)
                if isinstance(geom, (Polygon, MultiPolygon)):
                    if isinstance(geom, MultiPolygon):
                        polygons.extend(list(geom.geoms))
                    else:
                        polygons.append(geom)
                return polygons
            except Exception:
                return polygons
    return polygons

def build_neighborhood_index(cache_payload):
    data = cache_payload.get("data") if isinstance(cache_payload, dict) else cache_payload
    index = {}
    if not isinstance(data, list):
        return index
    for item in data:
        city_id = item.get("city_id")
        if city_id is None:
            continue
        name = item.get("name", "")
        polygons = normalize_polygon_from_item(item)
        if not polygons:
            continue
        entry = index.setdefault(city_id, {"polygons": [], "names": [], "centroids": []})
        for poly in polygons:
            try:
                poly_clean = poly.buffer(0) if not poly.is_valid else poly
            except Exception:
                poly_clean = poly
            c = poly_clean.centroid
            entry["polygons"].append(poly_clean)
            entry["names"].append(name)
            entry["centroids"].append((c.x, c.y))
    for cid, entry in index.items():
        pts = entry["centroids"]
        if pts:
            try:
                entry["kdtree"] = KDTree(pts)
            except Exception:
                entry["kdtree"] = None
        else:
            entry["kdtree"] = None
    return index

_NEIGH_CACHE = load_neighborhood_cache()
NEIGH_INDEX = build_neighborhood_index(_NEIGH_CACHE)

def identify_neighborhood(city_id, lat, lon, max_candidates=5):
    if city_id not in NEIGH_INDEX:
        return ""
    entry = NEIGH_INDEX[city_id]
    polygons = entry.get("polygons", [])
    names = entry.get("names", [])
    tree = entry.get("kdtree")
    if not polygons:
        return ""
    pt = Point(lon, lat)
    try:
        k = min(max_candidates, len(polygons))
        dists, idxs = tree.query([(lon, lat)], k=k)
        if np.isscalar(idxs):
            idxs = [int(idxs)]
        else:
            idxs = list(idxs[0])
    except Exception:
        idxs = list(range(min(max_candidates, len(polygons))))
    idxs = [int(i) for i in idxs if i is not None and i < len(polygons)]
    for i in idxs:
        poly = polygons[i]
        if poly.contains(pt):
            return names[i]
    for i in idxs:
        poly = polygons[i]
        if poly.intersects(pt):
            return names[i]
    for i in idxs:
        poly = polygons[i]
        if poly.buffer(1e-6).contains(pt):
            return names[i]
    for i, poly in enumerate(polygons):
        if poly.contains(pt):
            return names[i]
    return ""

# ---------------------
# fetch_collection
# ---------------------
async def fetch_collection(session, prop_id):
    url = URL_MATCH.format(prop_id)
    data = await fetch_json(session, url)
    if not data or not isinstance(data, list) or len(data) == 0:
        return ""
    melhor = max(data, key=lambda x: x.get("yield_boost", 1))
    if melhor.get("yield_boost", 1) >= 1.5:
        return melhor.get("name", "")
    return ""

# ---------------------
# fetch_property com delay aumentado
# ---------------------
async def fetch_property(session, prop_id, moeda, data_compra):
    url = URL_PROPERTY.format(prop_id)
    data = await fetch_json(session, url)
    if not data:
        return None
    
    # Delay entre requisições aumentado de 0.05 para 0.15
    await asyncio.sleep(0.15)

    preco = data.get("last_purchased_price", 0)
    rendimento = data.get("yield_per_hour", 0)
    colecao = await fetch_collection(session, prop_id)
    mint = round(rendimento * 176326.5459)

    if moeda == "USD":
        mint_usd = mint / 1000
        preco_display = preco / 1000
        markup = (preco_display / mint_usd) * 100 if mint_usd > 0 else 0
    else:
        preco_display = preco
        markup = (preco_display / mint) * 100 if mint > 0 else 0

    # construcao = ""
    # building = data.get("building")
    # if isinstance(building, dict):
    #     construcao = building.get("buildingName", "")
    # else:
    #     blds = data.get("buildings")
    #     if isinstance(blds, list) and len(blds) > 0 and isinstance(blds[0], dict):
    #         construcao = blds[0].get("buildingName", "")
    construcao = ""
    blds = data.get("buildings") 
    if isinstance(blds, list) and len(blds) > 0 and isinstance(blds[0], dict):
            construcao = blds[0].get("buildingName", "")
    else:
        building = data.get("buildings")
        if isinstance(building, dict):
            construcao = building.get("buildingName", "")

    try:
        lat_raw = data.get("centerlat")
        lon_raw = data.get("centerlng")
        lat = float(lat_raw) if lat_raw is not None else None
        lon = float(lon_raw) if lon_raw is not None else None
    except Exception:
        lat = None
        lon = None

    city_id = data.get("city", {}).get("id")

    bairro = ""
    if lat is not None and lon is not None and city_id is not None:
        bairro = identify_neighborhood(city_id, lat, lon)
        if not bairro:
            bairro = identify_neighborhood(city_id, lon, lat)
    else:
        bairro = ""

    return {
        "Data da Compra": data_compra,
        "ID Propriedade": int(prop_id),
        "Moeda": moeda,
        "Preço": float(preco_display),
        "Endereço": data.get("full_address"),
        "Cidade": data.get("city", {}).get("name"),
        "Bairro": bairro,
        "Proprietário": data.get("owner_username"),
        "Mint": int(mint),
        "Markup (%)": round(markup, 2),
        "Coleção": colecao,
        "Construção": construcao
    }

# ---------------------
# fetch_transactions com progresso
# ---------------------
async def fetch_transactions(url, moeda):
    async with aiohttp.ClientSession() as session:
        dados = await fetch_json(session, url)
        if not dados:
            print(f"❌ Erro ao buscar transações para {moeda}")
            return []

        actions = dados.get("actions", [])
        total = len(actions)
        print(f"📊 Processando {total} transações em {moeda}...")
        
        tasks = []
        for acao in actions:
            prop_id = acao.get("act", {}).get("data", {}).get("a45")
            if prop_id:
                timestamp = acao.get("@timestamp") or acao.get("timestamp")
                tasks.append(fetch_property(session, prop_id, moeda, timestamp))

        # Processa em lotes de 20 para evitar sobrecarga
        resultados = []
        batch_size = 35
        for i in range(0, len(tasks), batch_size):
            batch = tasks[i:i + batch_size]
            batch_results = await asyncio.gather(*batch, return_exceptions=True)
            resultados.extend([r for r in batch_results if r and not isinstance(r, Exception)])
            
            # Progresso
            progresso = min(i + batch_size, len(tasks))
            print(f"  ⏳ {progresso}/{len(tasks)} processadas...")
            
            # Pausa entre lotes
            if i + batch_size < len(tasks):
                await asyncio.sleep(2)
        
        print(f"  ✓ {len(resultados)} propriedades válidas em {moeda}")
        return resultados

# [... resto do código igual (main, formatação Excel, etc.) ...]
async def main():
    global request_count
    print("🚀 Iniciando busca de vendas...")
    print(f"⚙️  Configuração: Semáforo={SEM._value}, Delay=0.15s\n")
    
    print("📥 Buscando vendas UPX...")
    vendas_upx = await fetch_transactions(URL_N5, "UPX")

    print("\n⏸️  Aguardando 5 segundos antes das vendas USD...")
    await asyncio.sleep(5)

    print("📥 Buscando vendas USD...")
    vendas_usd = await fetch_transactions(URL_N52, "USD")

    todas_vendas = vendas_upx + vendas_usd
    if not todas_vendas:
        print("❌ Nenhuma venda encontrada.")
        return

    print(f"\n📊 Total de vendas coletadas: {len(todas_vendas)}")
    print(f"🌐 Total de requisições: {request_count}\n")

    df = pd.DataFrame(todas_vendas)

    df["Data da Compra"] = pd.to_datetime(df["Data da Compra"], errors="coerce", utc=True)
    df["Data"] = df["Data da Compra"].dt.tz_convert("America/Sao_Paulo").dt.strftime("%d/%m/%Y")
    df["Hora"] = df["Data da Compra"].dt.tz_convert("America/Sao_Paulo").dt.strftime("%H:%M:%S")

    colunas = [
        "ID Propriedade", "Data", "Hora", "Proprietário",
        "Preço", "Moeda", "Markup (%)", "Mint",
        "Endereço", "Cidade", "Bairro", "Coleção", "Construção"
    ]
    df = df[colunas]

    df["ID Propriedade"] = pd.to_numeric(df["ID Propriedade"], errors="coerce").astype("Int64")
    df["Preço"] = pd.to_numeric(df["Preço"], errors="coerce")
    df["Mint"] = pd.to_numeric(df["Mint"], errors="coerce").astype("Int64")

    df_upx = df[df["Moeda"] == "UPX"].copy()
    df_usd = df[df["Moeda"] == "USD"].copy()

    ordem_final = [
        "ID Propriedade", "Data", "Hora", "Proprietário",
        "Preço", "Moeda", "Markup (%)", "Mint",
        "Endereço", "Cidade", "Bairro", "Coleção", "Construção"
    ]

    df_upx = df_upx[ordem_final]
    df_usd = df_usd[ordem_final]

    cidade_upx = (
        df_upx.groupby("Cidade")
        .size()
        .reset_index(name="Quantidade de vendas (UPX)")
        .sort_values(by="Quantidade de vendas (UPX)", ascending=False)
    )

    cidade_usd = (
        df_usd.groupby("Cidade")
        .size()
        .reset_index(name="Quantidade de vendas (USD)")
        .sort_values(by="Quantidade de vendas (USD)", ascending=False)
    )

    bairro_upx = (
        df_upx[df_upx["Bairro"] != ""]
        .groupby("Bairro")
        .size()
        .reset_index(name="Quantidade de vendas (UPX)")
        .sort_values(by="Quantidade de vendas (UPX)", ascending=False)
        .head(35)
    )

    bairro_usd = (
        df_usd[df_usd["Bairro"] != ""]
        .groupby("Bairro")
        .size()
        .reset_index(name="Quantidade de vendas (USD)")
        .sort_values(by="Quantidade de vendas (USD)", ascending=False)
        .head(35)
    )

    print("📝 Criando planilha Excel...")
    # script: bot-upland/automacoes_adicionais/ultimas_vendas/
    # pasta:  bot-upland/relatorio_ultimas_vendas/ (dois níveis acima)
    pasta_saida = Path(__file__).parent.parent.parent / "relatorio_ultimas_vendas"
    pasta_saida.mkdir(exist_ok=True)
    arquivo_excel = str(pasta_saida / "vendas_upland.xlsx")

    with pd.ExcelWriter(arquivo_excel, engine="openpyxl") as writer:
        df_upx.to_excel(writer, sheet_name="UPX", index=False)
        df_usd.to_excel(writer, sheet_name="USD", index=False)

        col = 0
        cidade_upx.to_excel(writer, sheet_name="Estatísticas", index=False, startcol=col)
        col += len(cidade_upx.columns) + 2

        cidade_usd.to_excel(writer, sheet_name="Estatísticas", index=False, startcol=col)
        col += len(cidade_usd.columns) + 2

        bairro_upx.to_excel(writer, sheet_name="Estatísticas", index=False, startcol=col)
        col += len(bairro_upx.columns) + 2

        bairro_usd.to_excel(writer, sheet_name="Estatísticas", index=False, startcol=col)

    wb = load_workbook(arquivo_excel)

    from openpyxl.styles import Border, Side
    from openpyxl.utils import get_column_letter

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def aplicar_grade(ws):
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row,
                                min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = border

    def ajustar_largura(ws):
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max_len + 2

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        ws.freeze_panes = "A2"

        for col in range(1, ws.max_column + 1):
            header = ws.cell(1, col).value
            for c in ws.iter_cols(min_col=col, max_col=col, min_row=2):
                for cell in c:
                    if header == "ID Propriedade":
                        cell.number_format = "0"
                    elif header in ["Preço", "Mint"]:
                        if sheet_name == "USD" and header == "Preço":
                            cell.number_format = "#,##0.00"
                        else:
                            cell.number_format = "#,##0"
                    elif header == "Markup (%)":
                        cell.number_format = "0.00"

        aplicar_grade(ws)
        ajustar_largura(ws)

    wb.save(arquivo_excel)

    print("✅ Planilha 'vendas_upland.xlsx' criada com sucesso!")

if __name__ == "__main__":
    start = time.time()
    try:
        asyncio.run(main())
    except KeyboardInterrupt:
        print("\n⚠️ Interrompido pelo usuário.")
    finally:
        print(f"\n⏱️ Tempo total: {round(time.time() - start, 1)}s")